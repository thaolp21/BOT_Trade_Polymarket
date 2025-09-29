import re
import requests
import openpyxl
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import json
import os
import sys
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import time
import argparse
from datetime import timedelta

# Ensure project root (parent directory) is on sys.path for 'utils'
CURRENT_DIR = os.path.dirname(__file__)
PROJECT_ROOT = os.path.abspath(os.path.join(CURRENT_DIR, '..'))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from utils.common import r2, to_et_time, to_gmt7_date, to_gmt7_datetime, extract_time_part

load_dotenv()

# --- Config ---
DATA_API_ACTIVITY_URL = "https://data-api.polymarket.com/activity"
DATA_API_POSITION_URL = "https://data-api.polymarket.com/positions"
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")
POLYMARKET_ADDRESS = os.getenv("POLYMARKET_PROXY_ADDRESS")
FROM_TIME='1758727800' # Sep 24, 2025, 10:30:00 PM GMT+7
TO_TIME='1759251600'   # Oct 1, 2025, 12:00:00 AM GMT+7

# --- Analysis Configuration ---
# Price range to analyze (from high to low)
PRICE_RANGE = [5, 4, 3, 2, 1]

# Time frames for analysis (minutes from start)
TIME_FRAMES = [
    {'name': '0-4', 'min': 0, 'max': 4},
    {'name': '4-6', 'min': 4, 'max': 6},
    {'name': '6-8', 'min': 6, 'max': 8},
    {'name': '8-12', 'min': 8, 'max': 12}
]

# Column mapping for Excel Analysis sheet (starting from column 6)
# Each item maps to properties in price_obj and time_frames/win_time_frames
ANALYSIS_COLUMNS = [
    # Time frame columns (pairs: rounds + rate)
    {'type': 'time_frame', 'frame_index': 0, 'property': 'in_frame_rounds'},      
    {'type': 'time_frame', 'frame_index': 0, 'property': 'in_frame_rounds_rate'}, 
    {'type': 'time_frame', 'frame_index': 1, 'property': 'in_frame_rounds'},      
    {'type': 'time_frame', 'frame_index': 1, 'property': 'in_frame_rounds_rate'}, 
    {'type': 'time_frame', 'frame_index': 2, 'property': 'in_frame_rounds'},      
    {'type': 'time_frame', 'frame_index': 2, 'property': 'in_frame_rounds_rate'}, 
    # Win time frame columns (triplets: rounds + rate + ev)
    {'type': 'win_time_frame', 'frame_index': 0, 'property': 'win_in_frame_rounds'}, 
    {'type': 'win_time_frame', 'frame_index': 0, 'property': 'win_rate'},           
    {'type': 'win_time_frame', 'frame_index': 0, 'property': 'ev_value'},           
    {'type': 'win_time_frame', 'frame_index': 1, 'property': 'win_in_frame_rounds'}, 
    {'type': 'win_time_frame', 'frame_index': 1, 'property': 'win_rate'},           
    {'type': 'win_time_frame', 'frame_index': 1, 'property': 'ev_value'},           
    {'type': 'win_time_frame', 'frame_index': 2, 'property': 'win_in_frame_rounds'}, 
    {'type': 'win_time_frame', 'frame_index': 2, 'property': 'win_rate'},
    {'type': 'win_time_frame', 'frame_index': 2, 'property': 'ev_value'},
    ]
limit = 500
offset = 0 # max 10.000 for api: data-api.polymarket.com/positions / api data-api.polymarket.com/activity max 1.000
all_activities = []
#{'id': '0x6926ee8521e1d7528c725dbb023b2e4c05f4b7b2b488cf2a966a9e06d9db7318', 'title': 'Bitcoin Up or Down - September 18, 10:45AM-11:00AM ET', 'is_win': False, 'orders': [{'price': 10, 'time': 1758207034, 'time_to_matched': 5}], 'start_time': 1758206700},

# --- Helper: Convert timestamp to ET (Eastern Time) ---
def calculate_ev(win_rate, price):
    price_rr = 100/price if price != 0 else 0
    ev = win_rate * price_rr - 1
    return r2(ev)



# --- Analysis data ---
def get_analysis_data(grouped_data, from_time, to_time):
    from_time = int(from_time)
    to_time = int(to_time)
    # Each hour has 4 orders
    total_hours = (to_time - from_time) / 3600
    total_orders = int(total_hours * 4)
    total_trades = len(grouped_data)
    trade_order_rate = r2(total_trades / total_orders) if total_orders else 0
    win_rounds = [g for g in grouped_data if g.get('is_win')]
    total_win_rounds = len(win_rounds)
    win_trade_rate = r2(total_win_rounds / total_trades) if total_trades else 0
    analysis_prices = []
    for price in PRICE_RANGE:
        rounds_with_price = []
        for round_obj in grouped_data:
            # Check if any order in this round has this price
            for order in round_obj.get('orders', []):
                if order.get('price') == price:
                    # Attach time_to_matched for this price in this round
                    rounds_with_price.append({
                        'is_win': round_obj.get('is_win'),
                        'time_to_matched': order.get('time_to_matched')
                    })
                    break
        
        matched_rounds = len(rounds_with_price)
        matched_rate = r2(matched_rounds / total_trades) if total_trades else 0
        price_win_rounds = [r for r in rounds_with_price if r['is_win']]
        win_rate = r2(len(price_win_rounds) / matched_rounds) if matched_rounds else 0
        
        # Dynamic time frames
        time_frame_data = []
        win_time_frame_data = []
        
        for tf_config in TIME_FRAMES:
            tf_min, tf_max = tf_config['min'], tf_config['max']
            tf_name = tf_config['name']
            
            # Filter rounds by time frame
            tf_rounds = [r for r in rounds_with_price 
                        if r['time_to_matched'] is not None and tf_min <= r['time_to_matched'] < tf_max]
            win_tf_rounds = [r for r in price_win_rounds 
                           if r['time_to_matched'] is not None and tf_min <= r['time_to_matched'] < tf_max]
            
            # Time frame stats
            time_frame_data.append({
                'frame': tf_name,
                'in_frame_rounds': len(tf_rounds),
                'in_frame_rounds_rate': r2(len(tf_rounds)/matched_rounds) if matched_rounds else 0
            })
            
            # Win time frame stats
            tf_win_rate = r2(len(win_tf_rounds)/len(tf_rounds)) if len(tf_rounds) else 0
            tf_ev = calculate_ev(tf_win_rate, price)
            
            win_time_frame_data.append({
                'frame': tf_name,
                'win_in_frame_rounds': len(win_tf_rounds),
                'win_rate': tf_win_rate,
                'ev_value': tf_ev
            })
        
        # Overall price EV
        ev = calculate_ev(win_rate, price)

        analysis_prices.append({
            'price': price,
            'matched_rounds': matched_rounds,
            'matched_rate': matched_rate,
            'win_rounds': len(price_win_rounds),
            'win_rate': win_rate,
            'ev_value': r2(ev),
            'time_frames': time_frame_data,
            'win_time_frames': win_time_frame_data
        })
    return {
        'from_time': from_time,
        'to_time': to_time,
        'total_orders': total_orders,
        'total_trades': total_trades,
        'trade_order_rate': trade_order_rate,
        'win_rounds': total_win_rounds,
        'win_trade_rate': win_trade_rate,
        'analysis_prices': analysis_prices,
        'orders': grouped_data
    }

def parse_start_time_from_title(title):
    # Example: "Bitcoin Up or Down - September 18, 00:00AM-00:15AM ET" or "Bitcoin Up or Down - September 8, 9:00AM-9:15AM ET"
    # Extract date and start time
    match = re.search(r"- ([A-Za-z]+ \d+), (\d{1,2}:\d{2}[AP]M)-", title)
    if not match:
        return None
    date_str, time_str = match.groups()
    # Assume year is current year
    year = datetime.now().year
    # Compose datetime string
    dt_str = f"{date_str} {year} {time_str}"
    # Example: "September 18 2025 00:00AM"
    try:
        dt = datetime.strptime(dt_str, "%B %d %Y %I:%M%p")
        et = ZoneInfo('US/Eastern')
        dt = dt.replace(tzinfo=et)
        return int(dt.timestamp())
    except Exception:
        return None

def group_rounds(all_data):
    rounds = {}
    for item in all_data:
        cid = item.get('conditionId')
        if not cid:
            continue
        if cid not in rounds:
            rounds[cid] = {
                'id': cid,
                'title': item.get('title', ''),
                'is_win': item.get('is_win', False),
                'slug': item.get('slug', ''),
                'orders': []
            }
        # Always update is_win if any item in round is win
        if item.get('is_win'):
            rounds[cid]['is_win'] = True
        # Add order info
        order = {
            'price': int(round(item.get('price') * 100)),
            'time': item.get('timestamp')
        }
        rounds[cid]['orders'].append(order)

    # Post-process each round: unique orders by price, keep latest timestamp
    for r in rounds.values():
        price_map = {}
        for o in r['orders']:
            p = o['price']
            t = o['time']
            if p not in price_map or (t is not None and t > price_map[p]['time']):
                price_map[p] = o
        # Now, price_map.values() are unique by price, latest timestamp
        r['orders'] = list(price_map.values())
        # Sort orders by price descending
        r['orders'].sort(key=lambda x: (-x['price'], x['time'] if x['time'] is not None else 0))
        # Parse start_time from title
        r['start_time'] = parse_start_time_from_title(r['title'])
        # Add time_to_matched for each order
        for o in r['orders']:
            if r['start_time'] is not None and o['time'] is not None:
                o['time_to_matched'] = int((o['time'] - r['start_time']) / 60)
            else:
                o['time_to_matched'] = None
    # Return as list
    return list(rounds.values())

# --- Get data from API or mock file ---

def get_data():
    all_data = []
    redeemed_orders = []
    try:
        start_time = int(FROM_TIME)
        now_time = int(TO_TIME) if TO_TIME != "" else int(datetime.now(timezone.utc).timestamp())
        three_hours = 3 * 60 * 60
        while start_time < now_time:
            end_time = min(start_time + three_hours, now_time)
            params = {
                'user': POLYMARKET_ADDRESS,
                'start': start_time,
                'end': end_time,
                'limit': 500,
            }
            try:
                resp = requests.get(DATA_API_ACTIVITY_URL, params=params, timeout=10)
                resp.raise_for_status()
                batch = resp.json()
                if isinstance(batch, list):
                    # Filter and split by type
                    for item in batch:
                        if item.get('type') == 'TRADE':
                            all_data.append(item)
                        elif item.get('type') == 'REDEEM':
                            redeemed_orders.append(item)
            except Exception as e:
                print(f"Warning: API error for window {start_time} - {end_time}: {e}")
            start_time = end_time
            time.sleep(0.2)  # avoid hammering API

        # Fetch extended info from positions API with pagination
        try:
            user_addr = POLYMARKET_ADDRESS
            limit = 500
            max_offset = 10000
            offset = 0
            all_positions = []
            while offset < max_offset:
                params = {
                    'user': user_addr,
                    'limit': limit,
                    'offset': offset,
                    'sortBy': 'CURRENT',
                    'sortDirection': 'DESC'
                }
                resp = requests.get(DATA_API_POSITION_URL, params=params, timeout=10)
                resp.raise_for_status()
                page = resp.json()
                if isinstance(page, dict) and 'data' in page:
                    page = page['data']
                if not page:
                    break
                all_positions.extend(page)
                if len(page) < limit:
                    break
                offset += limit
                time.sleep(0.2)

            redeemable_map = {}
            for pos in all_positions:
                if pos.get('redeemable') == True:
                    cid = pos.get('conditionId')
                    if cid:
                        redeemable_map[cid] = {
                            'totalBought': pos.get('totalBought'),
                            'avgPrice': pos.get('avgPrice'),
                            'cashPnl': pos.get('cashPnl'),
                            'percentPnl': pos.get('percentPnl')
                        }
            # Mark is_win from redeemed_orders
            redeemed_cids = set()
            for r in redeemed_orders:
                cid = r.get('conditionId')
                if cid:
                    redeemed_cids.add(cid)
            for obj in all_data:
                cid = obj.get('conditionId')
                if cid in redeemed_cids:
                    obj['is_win'] = True
                # Merge info into data
                if cid in redeemable_map:
                    obj.update(redeemable_map[cid])
                    # If percentPnl > 0, set is_win True
                    percent_pnl = obj.get('percentPnl')
                    if percent_pnl is not None and percent_pnl > 0:
                        obj['is_win'] = True
        except Exception as e:
            print(f"Warning: Could not fetch/merge positions API: {e}")
        return all_data
    except Exception as e:
        # fallback to mock
        print(f"Warning: Could not fetch data from API. Error: {e}")

# --- Main fill function ---

def fill_excel_analysis_sheet(wb, analysis_data):
    """
    Fill the summary section (A2:C6) in the 'Analysis' sheet with:
    From, To, Total orders, Total trades - trade/order rate, Win - win/trade rate
    """
    try:
        analysis_prices = analysis_data.get('analysis_prices', [])
        if 'Analysis' in wb.sheetnames:
            ws = wb['Analysis']
            start_row = 11
            ws['B2'] = to_et_time(analysis_data.get('from_time')) if analysis_data.get('from_time') else ''
            ws['B3'] = to_et_time(analysis_data.get('to_time')) if analysis_data.get('to_time') else ''
            ws['B4'] = analysis_data.get('total_orders')
            ws['B5'] = analysis_data.get('total_trades')
            ws['C5'] = analysis_data.get('trade_order_rate')
            ws['B6'] = analysis_data.get('win_rounds')
            ws['C6'] = analysis_data.get('win_trade_rate')

            for i, price_obj in enumerate(analysis_prices):
                row = start_row + i
                # Fixed columns (1-6)
                ws.cell(row=row, column=1, value=price_obj['price'])
                ws.cell(row=row, column=2, value=price_obj['matched_rounds'])
                ws.cell(row=row, column=3, value=price_obj['win_rounds'])
                ws.cell(row=row, column=4, value=price_obj['win_rate'])
                ws.cell(row=row, column=5, value=price_obj['ev_value'])
                
                # Dynamic columns (7+) based on ANALYSIS_COLUMNS configuration
                tf = price_obj.get('time_frames', [])
                win_tf = price_obj.get('win_time_frames', [])
                
                for col_idx, col_config in enumerate(ANALYSIS_COLUMNS):
                    column_num = 6 + col_idx  # Start from column 6
                    value = None
                    
                    if col_config['type'] == 'time_frame':
                        frame_idx = col_config['frame_index']
                        if frame_idx < len(tf):
                            value = tf[frame_idx].get(col_config['property'])
                    elif col_config['type'] == 'win_time_frame':
                        frame_idx = col_config['frame_index']
                        if frame_idx < len(win_tf):
                            value = win_tf[frame_idx].get(col_config['property'])
                    
                    ws.cell(row=row, column=column_num, value=value)
    except Exception as e:
        print(f"Warning: Could not fill analysis summary section: {e}")

def _build_report_filename(from_ts: int, to_ts: int) -> str:
    """Return filename like dd.MM.yyyy-dd.MM.yyyy.xlsx using GMT+7.
    to_ts treated as exclusive end; subtract 1 second for inclusive date.
    """
    from_date = to_gmt7_date(int(from_ts))
    to_inclusive = int(to_ts) - 1 if to_ts else int(time.time())
    to_date = to_gmt7_date(to_inclusive)
    return f"{from_date}-{to_date}.xlsx"


def fill_excel(data):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    effective_to = TO_TIME if TO_TIME else int(datetime.now(timezone.utc).timestamp())
    analysis_data = get_analysis_data(data, FROM_TIME, effective_to)
    fill_excel_analysis_sheet(wb, analysis_data)
    ws = wb['Data']
    start_row = 2  # assuming headers are in row 1
    row = start_row
    for idx, obj in enumerate(data, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=obj.get('title', ''))
        ws.cell(row=row, column=10, value=obj.get('id', ''))
        ws.cell(row=row, column=11, value=obj.get('slug', ''))
        ws.cell(row=row, column=5, value='WIN' if obj.get('is_win') else 'LOSE')
        cell = ws.cell(row=row, column=5)
        if obj.get('is_win'):
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for order in obj.get('orders', []):
            ts = order.get('time')
            ws.cell(row=row, column=3, value=extract_time_part(ts))
            ws.cell(row=row, column=4, value=order.get('price'))
            ws.cell(row=row, column=6, value=order.get('time_to_matched'))
            ws.cell(row=row, column=12, value=ts)
            row += 1
    filename = _build_report_filename(FROM_TIME, effective_to)
    output_path = os.path.join(os.path.dirname(__file__), filename)
    wb.save(output_path)
    # Write metadata for other scripts
    meta = {
        'path': output_path,
        'filename': filename,
        'from_time': FROM_TIME,
        'to_time': effective_to
    }
    try:
        with open(os.path.join(os.path.dirname(__file__), 'last_report.json'), 'w', encoding='utf-8') as f:
            json.dump(meta, f, indent=2)
    except Exception as e:
        print(f"Warning: could not write metadata file: {e}")
    print(f"Filled data saved to {output_path}")
    return output_path

def main():
    if not POLYMARKET_ADDRESS:
        print("POLYMARKET_PROXY_ADDRESS not set in .env; cannot fetch data. Exiting.")
        exit(1)

    from_dt = to_gmt7_datetime(int(FROM_TIME))
    to_dt = to_gmt7_datetime(int(TO_TIME))
    print(f"Data range: {from_dt} to {to_dt} [GMT+7]")
    
    # Fetch and process data
    data = get_data()
    grouped_data = group_rounds(data)
    print(f"Total unique rounds: {len(grouped_data)}")
    
    # Generate Excel report
    generated_path = fill_excel(grouped_data)
    print(f"Report generated: {generated_path}")
    
if __name__ == "__main__":
    main()   
